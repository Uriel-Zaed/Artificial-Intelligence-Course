
from problem import Problem
from branch_and_bound import BranchAndBound
from branch_and_bound_FC import BranchAndBound_FC
from branch_and_bound_hur import BranchAndBound_Hur

import matplotlib.pyplot as plt
from openpyxl import Workbook

N = 8
D = 8
num_of_problems = 10
p1 = [0.4, 0.7]
p2 = [0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9]
problems = []

graphs = {}

for p_1 in p1:
    graph_p2 = []
    graph_cc_bnb = []
    graph_cc_bnb_fc = []
    graph_cc_bnb_hur = []

    for p_2 in p2:
        print(f"p1={p_1}, p2={p_2} is calculating...")
        average_CC_bnb = 0
        average_CC_bnb_fc = 0
        average_CC_bnb_hur = 0
        for _ in range(num_of_problems):
            
            pr = Problem(N,D,p_1,p_2)
            bnb = BranchAndBound(pr)
            bnb.solver()
            average_CC_bnb += pr.CC/num_of_problems

            pr.CC_reset()
            bnb_fc = BranchAndBound_FC(pr)
            bnb_fc.solver()
            average_CC_bnb_fc += pr.CC/num_of_problems

            pr.CC_reset()
            bnb_hur = BranchAndBound_Hur(pr)
            bnb_hur.solver()
            average_CC_bnb_hur += pr.CC/num_of_problems

        graph_p2.append(p_2)
        graph_cc_bnb.append(average_CC_bnb)
        graph_cc_bnb_fc.append(average_CC_bnb_fc)
        graph_cc_bnb_hur.append(average_CC_bnb_hur)
        print(f"p1={p_1}, p2={p_2} is done\n")
    graphs[p_1] = (graph_p2, graph_cc_bnb, graph_cc_bnb_fc, graph_cc_bnb_hur)


# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Write data to the worksheet
for row_idx, row_data in enumerate(graphs[0.4], start=1):
    for col_idx, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=cell_data)

# Save the workbook
wb.save(f"data p1=0.4_N={N}_D={D}.xlsx")

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Write data to the worksheet
for row_idx, row_data in enumerate(graphs[0.7], start=1):
    for col_idx, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=cell_data)

# Save the workbook
wb.save(f"data p1=0.7_N={N}_D={D}.xlsx")
        
for p_1 in graphs:
    plt.plot(graphs[p_1][0], graphs[p_1][1], 'o-', label='bnb')
    plt.plot(graphs[p_1][0], graphs[p_1][2], 'o-', label='bnb_fc')
    plt.plot(graphs[p_1][0], graphs[p_1][3], 'o-', label='bnb_hur')
    plt.legend()
    plt.xlabel("p2")
    plt.ylabel("averege CC")
    plt.title(f"B&B - p1 = {p_1}")
    plt.savefig(f'p1={p_1}_N={N}_D={D}.png')
    plt.show()
